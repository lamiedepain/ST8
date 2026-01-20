#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fonctions utilitaires pour l'application ST8 Planning
Inclut: Excel, SharePoint, backups, helpers
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
import requests
from requests_ntlm import HttpNtlmAuth
import openpyxl
import msal

from config import (
    EXCEL_FILE, LOCAL_EXCEL_PATH, BACKUP_DIR, MAX_BACKUPS,
    SHAREPOINT_FILE_URL, SHAREPOINT_USERNAME, SHAREPOINT_PASSWORD,
    SHAREPOINT_SITE_URL, SHAREPOINT_CLIENT_ID, SHAREPOINT_CLIENT_SECRET,
    SHAREPOINT_TENANT_ID, AUTO_SYNC_ENABLED, SHAREPOINT_SITE_NAME,
    SHAREPOINT_DRIVE_PATH, SHAREPOINT_FILE_NAME, SHAREPOINT_FILE_PATH
)

# ==============================================================================
# EXCEPTIONS PERSONNALISÉES
# ==============================================================================

class ExcelError(Exception):
    """Exception levée pour les erreurs liées au fichier Excel"""
    pass

class SharePointError(Exception):
    """Exception levée pour les erreurs liées à SharePoint"""
    pass

# ==============================================================================
# FONCTIONS SHAREPOINT
# ==============================================================================

def download_from_sharepoint(dest_path=None):
    """
    Télécharge le fichier Excel depuis SharePoint via Microsoft Graph API
    
    Args:
        dest_path: Chemin de destination (défaut: LOCAL_EXCEL_PATH)
    
    Returns:
        Path: Chemin du fichier téléchargé
    
    Raises:
        SharePointError: Si le téléchargement échoue
    """
    if dest_path is None:
        dest_path = LOCAL_EXCEL_PATH
    
    print(f"📥 Téléchargement depuis SharePoint...")
    
    try:
        # Priorité à OAuth2 avec Microsoft Graph (méthode moderne)
        if SHAREPOINT_CLIENT_ID and SHAREPOINT_CLIENT_SECRET:
            print(f"🔐 Utilisation de Microsoft Graph API...")
            token = get_sharepoint_access_token()
            
            # Construire l'URL Microsoft Graph
            # Format: /sites/{site-name}/drive/root:/{path-to-file}:/content
            graph_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_NAME}/drive/root:/{SHAREPOINT_DRIVE_PATH}/{SHAREPOINT_FILE_NAME}:/content"
            
            print(f"🔗 Téléchargement: {SHAREPOINT_FILE_NAME}")
            
            response = requests.get(
                graph_url,
                headers={'Authorization': f'Bearer {token}'},
                timeout=120
            )
            
            if response.status_code == 200:
                with open(dest_path, 'wb') as f:
                    f.write(response.content)
                file_size = len(response.content) / 1024 / 1024
                print(f"✅ Fichier téléchargé: {dest_path} ({file_size:.2f} MB)")
                return Path(dest_path)
            else:
                error_msg = f"Échec téléchargement Graph API: {response.status_code}"
                try:
                    error_detail = response.json()
                    error_msg += f" - {error_detail.get('error', {}).get('message', '')}"
                except:
                    pass
                raise SharePointError(error_msg)
        
        # Fallback: NTLM (moins recommandé pour SharePoint Online)
        elif SHAREPOINT_USERNAME and SHAREPOINT_PASSWORD:
            print(f"⚠️  Tentative NTLM (peut ne pas fonctionner avec SharePoint Online)...")
            auth = HttpNtlmAuth(SHAREPOINT_USERNAME, SHAREPOINT_PASSWORD)
            
            download_url = f"{SHAREPOINT_SITE_URL}/_layouts/15/download.aspx?SourceUrl={SHAREPOINT_FILE_PATH}"
            
            response = requests.get(
                download_url,
                auth=auth,
                allow_redirects=True,
                timeout=60
            )
            
            if response.status_code == 200:
                with open(dest_path, 'wb') as f:
                    f.write(response.content)
                print(f"✅ Fichier téléchargé: {dest_path}")
                return Path(dest_path)
            else:
                raise SharePointError(f"Échec NTLM: {response.status_code} - Configurez OAuth2 (CLIENT_ID et CLIENT_SECRET) dans .env")
        
        else:
            raise SharePointError("❌ Credentials SharePoint manquants. Configurez CLIENT_ID et CLIENT_SECRET dans le fichier .env")
    
    except SharePointError:
        raise
    except Exception as e:
        raise SharePointError(f"Erreur téléchargement: {str(e)}")

def get_sharepoint_access_token():
    """
    Obtient un token d'accès OAuth2 pour SharePoint via Microsoft Graph
    
    Returns:
        str: Access token
    
    Raises:
        SharePointError: Si l'authentification échoue
    """
    try:
        # Configuration MSAL
        authority = f"https://login.microsoftonline.com/{SHAREPOINT_TENANT_ID}"
        scope = ["https://graph.microsoft.com/.default"]
        
        print(f"🔐 Authentification Azure AD...")
        print(f"   Tenant ID: {SHAREPOINT_TENANT_ID}")
        print(f"   Client ID: {SHAREPOINT_CLIENT_ID[:8]}...")
        
        # Créer l'application confidentielle
        app = msal.ConfidentialClientApplication(
            SHAREPOINT_CLIENT_ID,
            authority=authority,
            client_credential=SHAREPOINT_CLIENT_SECRET,
        )
        
        # Acquérir le token
        result = app.acquire_token_for_client(scopes=scope)
        
        if "access_token" in result:
            print(f"✅ Token obtenu")
            return result["access_token"]
        else:
            error = result.get("error_description", result.get("error", "Erreur inconnue"))
            raise SharePointError(f"Échec authentification Azure AD: {error}")
            
    except Exception as e:
        raise SharePointError(f"Erreur authentification: {str(e)}")
    except Exception as e:
        raise SharePointError(f"Erreur authentification SharePoint: {str(e)}")

def upload_to_sharepoint(source_path=None):
    """
    Upload le fichier Excel vers SharePoint (backup cloud)
    
    Args:
        source_path: Chemin source (défaut: LOCAL_EXCEL_PATH)
    
    Returns:
        bool: True si succès
    
    Raises:
        SharePointError: Si l'upload échoue
    """
    if source_path is None:
        source_path = LOCAL_EXCEL_PATH
    
    print(f"📤 Upload vers SharePoint...")
    
    try:
        if SHAREPOINT_CLIENT_ID and SHAREPOINT_CLIENT_SECRET:
            print(f"🔐 Utilisation de Microsoft Graph API...")
            token = get_sharepoint_access_token()
            
            # URL pour uploader via Graph API
            graph_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_NAME}/drive/root:/{SHAREPOINT_DRIVE_PATH}/{SHAREPOINT_FILE_NAME}:/content"
            
            with open(source_path, 'rb') as f:
                response = requests.put(
                    graph_url,
                    headers={
                        'Authorization': f'Bearer {token}',
                        'Content-Type': 'application/octet-stream'
                    },
                    data=f,
                    timeout=120
                )
            
            if response.status_code in [200, 201]:
                print(f"✅ Fichier uploadé vers SharePoint")
                return True
            else:
                error_msg = f"Échec upload: {response.status_code}"
                try:
                    error_detail = response.json()
                    error_msg += f" - {error_detail.get('error', {}).get('message', '')}"
                except:
                    pass
                raise SharePointError(error_msg)
        
        elif SHAREPOINT_USERNAME and SHAREPOINT_PASSWORD:
            raise SharePointError("❌ Authentification NTLM non supportée. Configurez OAuth2 (CLIENT_ID et CLIENT_SECRET)")
        
        else:
            raise SharePointError("❌ Credentials SharePoint manquants")
    
    except SharePointError:
        raise
    except Exception as e:
        raise SharePointError(f"Erreur upload: {str(e)}")

def sync_from_sharepoint_if_enabled():
    """
    Synchronise depuis SharePoint si AUTO_SYNC_ENABLED = True
    
    Returns:
        bool: True si synchronisé, False si désactivé
    """
    if not AUTO_SYNC_ENABLED:
        print("ℹ️  Synchronisation SharePoint désactivée")
        return False
    
    try:
        # Créer backup avant sync
        if LOCAL_EXCEL_PATH.exists():
            create_backup(LOCAL_EXCEL_PATH)
        
        # Télécharger depuis SharePoint
        download_from_sharepoint()
        return True
    
    except SharePointError as e:
        print(f"⚠️  Synchronisation échouée: {e}")
        print(f"ℹ️  Utilisation du fichier local: {LOCAL_EXCEL_PATH}")
        return False

# ==============================================================================
# FONCTIONS EXCEL
# ==============================================================================

def load_workbook(excel_path=None, data_only=True):
    """
    Charge un classeur Excel
    
    Args:
        excel_path: Chemin du fichier (défaut: LOCAL_EXCEL_PATH)
        data_only: Si True, charge les valeurs (pas les formules)
    
    Returns:
        openpyxl.Workbook: Classeur chargé
    
    Raises:
        ExcelError: Si le chargement échoue
    """
    if excel_path is None:
        excel_path = LOCAL_EXCEL_PATH
    
    if not Path(excel_path).exists():
        raise ExcelError(f"Fichier Excel introuvable: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=data_only, keep_vba=True)
        return wb
    except Exception as e:
        raise ExcelError(f"Erreur chargement Excel: {str(e)}")

def save_workbook(wb, dest_path=None):
    """
    Sauvegarde un classeur Excel
    
    Args:
        wb: Classeur à sauvegarder
        dest_path: Chemin de destination (défaut: LOCAL_EXCEL_PATH)
    
    Raises:
        ExcelError: Si la sauvegarde échoue
    """
    if dest_path is None:
        dest_path = LOCAL_EXCEL_PATH
    
    try:
        wb.save(dest_path)
        print(f"✅ Fichier sauvegardé: {dest_path}")
    except Exception as e:
        raise ExcelError(f"Erreur sauvegarde Excel: {str(e)}")

def cell_to_str(value):
    """
    Convertit une valeur de cellule Excel en chaîne propre
    
    Args:
        value: Valeur de la cellule
    
    Returns:
        str: Chaîne nettoyée (ou None)
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()

def cell_to_date_str(value):
    """
    Convertit une valeur de cellule Excel en date formatée (sans heure)
    
    Args:
        value: Valeur de la cellule (datetime, str, etc.)
    
    Returns:
        str: Date formatée JJ/MM/AAAA ou chaîne originale
    """
    if value is None:
        return None
    
    # Si c'est un datetime, formater sans l'heure
    if isinstance(value, datetime):
        # Filtrer les dates invalides (avant 1950)
        if value.year < 1950:
            return None
        return value.strftime('%d/%m/%Y')
    
    # Si c'est une string qui ressemble à une date avec heure
    if isinstance(value, str):
        value = value.strip()
        # Essayer de parser et reformater
        try:
            # Format avec heure : "2026-01-15 00:00:00"
            if ' ' in value and ':' in value:
                dt = datetime.strptime(value.split()[0], '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y')
        except:
            pass
        return value
    
    return str(value).strip()

def normalize_status(value):
    """
    Normalise un code de statut (présence/absence)
    
    Args:
        value: Code brut
    
    Returns:
        str: Code normalisé en majuscules
    """
    if value is None or value == '':
        return 'P'  # Présent par défaut
    
    status = str(value).strip().upper()
    return status if status else 'P'

def find_header_row(sheet, search_value='Matricule'):
    """
    Trouve la ligne d'en-tête contenant une valeur spécifique
    
    Args:
        sheet: Feuille Excel
        search_value: Valeur à chercher
    
    Returns:
        int: Numéro de ligne (1-indexed) ou None
    """
    for row_idx in range(1, min(sheet.max_row, 20) + 1):
        for cell in sheet[row_idx]:
            if cell.value and search_value.lower() in str(cell.value).lower():
                return row_idx
    return None

# ==============================================================================
# FONCTIONS BACKUP
# ==============================================================================

def init_backup_dir():
    """Crée le dossier de backups s'il n'existe pas"""
    BACKUP_DIR.mkdir(exist_ok=True)

def create_backup(source_file=None):
    """
    Crée une sauvegarde horodatée du fichier Excel
    
    Args:
        source_file: Fichier source (défaut: LOCAL_EXCEL_PATH)
    
    Returns:
        Path: Chemin du fichier de backup créé
    
    Raises:
        ExcelError: Si la création échoue
    """
    if source_file is None:
        source_file = LOCAL_EXCEL_PATH
    
    source_path = Path(source_file)
    
    if not source_path.exists():
        raise ExcelError(f"Fichier source introuvable: {source_file}")
    
    # Créer dossier backups
    init_backup_dir()
    
    # Générer nom avec timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"backup_{timestamp}{source_path.suffix}"
    backup_path = BACKUP_DIR / backup_name
    
    # Copier fichier
    try:
        shutil.copy2(source_path, backup_path)
        print(f"✅ Backup créé: {backup_path}")
        
        # Nettoyer anciens backups
        cleanup_old_backups()
        
        return backup_path
    
    except Exception as e:
        raise ExcelError(f"Erreur création backup: {str(e)}")

def cleanup_old_backups():
    """Supprime les backups excédentaires (garde les MAX_BACKUPS plus récents)"""
    if not BACKUP_DIR.exists():
        return
    
    backups = sorted(BACKUP_DIR.glob('backup_*.xlsm'), key=lambda p: p.stat().st_mtime, reverse=True)
    
    if len(backups) > MAX_BACKUPS:
        for old_backup in backups[MAX_BACKUPS:]:
            try:
                old_backup.unlink()
                print(f"🗑️  Backup supprimé: {old_backup.name}")
            except Exception as e:
                print(f"⚠️  Erreur suppression {old_backup.name}: {e}")

def list_backups():
    """
    Liste tous les backups disponibles
    
    Returns:
        list: Liste de tuples (nom, date, taille)
    """
    if not BACKUP_DIR.exists():
        return []
    
    backups = []
    for backup_file in sorted(BACKUP_DIR.glob('backup_*.xlsm'), key=lambda p: p.stat().st_mtime, reverse=True):
        stat = backup_file.stat()
        backups.append({
            'name': backup_file.name,
            'path': str(backup_file),
            'date': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'size': stat.st_size
        })
    
    return backups

def restore_backup(backup_name):
    """
    Restaure un backup spécifique
    
    Args:
        backup_name: Nom du fichier de backup
    
    Returns:
        Path: Chemin du fichier restauré
    
    Raises:
        ExcelError: Si la restauration échoue
    """
    backup_path = BACKUP_DIR / backup_name
    
    if not backup_path.exists():
        raise ExcelError(f"Backup introuvable: {backup_name}")
    
    # Créer backup du fichier actuel avant restauration
    if LOCAL_EXCEL_PATH.exists():
        create_backup(LOCAL_EXCEL_PATH)
    
    # Restaurer
    try:
        shutil.copy2(backup_path, LOCAL_EXCEL_PATH)
        print(f"✅ Backup restauré: {backup_name}")
        return LOCAL_EXCEL_PATH
    except Exception as e:
        raise ExcelError(f"Erreur restauration backup: {str(e)}")

# ==============================================================================
# FONCTIONS FICHIERS
# ==============================================================================

def list_excel_files(directory=None):
    """
    Liste tous les fichiers Excel dans un répertoire
    
    Args:
        directory: Répertoire à scanner (défaut: dossier du script)
    
    Returns:
        list: Liste de fichiers Excel (.xlsx, .xlsm)
    """
    if directory is None:
        directory = Path(__file__).parent
    
    directory = Path(directory)
    excel_files = []
    
    for ext in ['*.xlsx', '*.xlsm']:
        excel_files.extend(directory.glob(ext))
    
    return [{'name': f.name, 'path': str(f), 'size': f.stat().st_size} for f in sorted(excel_files)]

def resolve_excel_path(file_param=None):
    """
    Résout le chemin d'un fichier Excel
    
    Args:
        file_param: Nom ou chemin de fichier (défaut: EXCEL_FILE)
    
    Returns:
        Path: Chemin absolu du fichier
    
    Raises:
        ExcelError: Si le fichier n'existe pas
    """
    if file_param is None:
        return LOCAL_EXCEL_PATH
    
    # Chemin absolu
    path = Path(file_param)
    if path.is_absolute() and path.exists():
        return path
    
    # Relatif au dossier script
    path = Path(__file__).parent / file_param
    if path.exists():
        return path
    
    raise ExcelError(f"Fichier Excel introuvable: {file_param}")

# ==============================================================================
# VALIDATION
# ==============================================================================

def validate_excel_structure(excel_path=None):
    """
    Valide la structure du fichier Excel
    
    Args:
        excel_path: Chemin du fichier (défaut: LOCAL_EXCEL_PATH)
    
    Returns:
        dict: Résultat de la validation
    
    Raises:
        ExcelError: Si la structure est invalide
    """
    if excel_path is None:
        excel_path = LOCAL_EXCEL_PATH
    
    wb = load_workbook(excel_path)
    
    results = {
        'valid': True,
        'errors': [],
        'warnings': [],
        'sheets': wb.sheetnames,
        'agents_count': 0
    }
    
    # Vérifier feuille 'config'
    if 'config' not in wb.sheetnames:
        results['valid'] = False
        results['errors'].append("Feuille 'config' manquante")
    else:
        config_sheet = wb['config']
        header_row = find_header_row(config_sheet, 'Matricule')
        if not header_row:
            results['warnings'].append("En-tête 'Matricule' introuvable dans 'config'")
        else:
            # Compter agents
            agents_count = 0
            for row_idx in range(header_row + 1, config_sheet.max_row + 1):
                matricule = config_sheet.cell(row_idx, 1).value
                if matricule:
                    agents_count += 1
            results['agents_count'] = agents_count
    
    return results

if __name__ == "__main__":
    print("Utilitaires ST8 Planning")
    print("=" * 60)
    
    # Test connexion SharePoint
    print("\n1. Test SharePoint:")
    try:
        if AUTO_SYNC_ENABLED:
            sync_from_sharepoint_if_enabled()
        else:
            print("   Synchronisation désactivée")
    except Exception as e:
        print(f"   ⚠️  Erreur: {e}")
    
    # Test chargement Excel
    print("\n2. Test Excel local:")
    try:
        wb = load_workbook()
        print(f"   ✅ Fichier chargé: {len(wb.sheetnames)} feuilles")
    except Exception as e:
        print(f"   ⚠️  Erreur: {e}")
    
    # Test backups
    print("\n3. Backups disponibles:")
    backups = list_backups()
    print(f"   {len(backups)} backup(s) trouvé(s)")
    for b in backups[:3]:
        print(f"   - {b['name']} ({b['date']})")
